import argparse
import csv
import json
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

import requests

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from AI_analysis_codex_on_ticker import (  # noqa: E402
    build_evidence_pack,
    call_codex_cli,
    write_report,
)
from download_news_links_working import (  # noqa: E402
    build_query,
    get_news_items,
    load_company_names,
    write_outputs as write_news_outputs,
)
from download_screener_transcripts import (  # noqa: E402
    clear_old_concall_dirs,
    download_pdf as download_transcript_pdf,
    extract_transcripts,
    filename_from_url,
    load_screener_html,
    save_metadata as save_transcript_metadata,
)
from filter_investment_reports_on_score import (  # noqa: E402
    METRICS as SCORE_METRICS,
    format_score,
    parse_scores,
)


DATA_DIR = ROOT_DIR / "data"
DEFAULT_CSV_FILE = ROOT_DIR / "config" / "nifty100.csv"
PROMPT_FILE = ROOT_DIR / "config" / "investment_analysis_prompt.md"

MANUAL_SYMBOLS = [
    "LT",
    "ANANTRAJ",
    "CUMMINSIND",
    "MINDSPACE",
    "NTPC",
    "GVT&D",
    "NETWEB",
    "ABB",
    "VOLTAS",
    "BLUESTARCO",
    "HAL",
    "ZENTEC",
    "DATAPATTNS",
    "HBLENGINE",
    "TATAPOWER",
    "WAAREEENER",
    "PERSISTENT",
    "COFORGE",
    "AFFLE",
    "KPITTECH",
    "MOTHERSON",
    "HAVELLS",
    "ETERNAL",
    "OLAELEC",
    "KAYNES",
    "DIXON",
    "SYRMA",
    "MOSCHIP",
    "IKS",
    "SONACOMS",
    "POLYMED",
    "TRITURBINE",
    "NAVINFLUOR",
    "NEULANDLAB",
    "CAPLIPOINT",
    "AEGISLOG",
    "RRKABEL",
    "FINCABLES",
    "ELGIEQUIP",
    "CCL",
    "REDINGTON",
    "EIHOTEL",
]

MAX_ANNUAL_REPORTS = 5
MAX_TRANSCRIPTS = 10
MAX_NEWS_LINKS = 15

NSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X) "
        "AppleWebKit/537.36 Chrome/125 Safari/537.36"
    ),
    "Accept": "application/json,text/plain,*/*",
    "Referer": "https://www.nseindia.com/",
}

SCREENER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/148.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def normalize_symbol(symbol):
    return (symbol or "").strip().upper()


def normalize_xlsx_symbol(symbol):
    return re.sub(r"[^A-Za-z]", "", str(symbol or "").strip()).upper()


def parse_symbol_items(items):
    symbols = []

    for item in items or []:
        for symbol in item.split(","):
            symbol = normalize_symbol(symbol)
            if symbol and symbol not in symbols:
                symbols.append(symbol)

    return symbols


def load_symbols_from_csv(csv_file):
    csv_path = Path(csv_file)

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    symbols = []

    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)

        if "SYMBOL" not in (reader.fieldnames or []):
            raise ValueError("CSV must contain SYMBOL column.")

        for row in reader:
            symbol = normalize_symbol(row.get("SYMBOL"))
            if symbol and symbol not in symbols:
                symbols.append(symbol)

    return symbols


def load_symbols_from_xlsx(xlsx_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required to read .xlsx ticker files.") from exc

    xlsx_path = Path(xlsx_file)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)

    if not header_row:
        raise ValueError("XLSX file must contain a header row.")

    ticker_column = None
    for index, header in enumerate(header_row):
        if str(header or "").strip().lower() == "ticker":
            ticker_column = index
            break

    if ticker_column is None:
        raise ValueError("XLSX file must contain Ticker column.")

    symbols = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[ticker_column] if ticker_column < len(row) else None
        symbol = normalize_xlsx_symbol(value)
        if symbol and symbol not in symbols:
            symbols.append(symbol)

    workbook.close()
    return symbols


def load_symbols_from_file(symbol_file):
    symbol_path = Path(symbol_file)
    if symbol_path.suffix.lower() == ".xlsx":
        return load_symbols_from_xlsx(symbol_path)

    return load_symbols_from_csv(symbol_path)


def load_company_names_from_file(symbol_file):
    symbol_path = Path(symbol_file)
    if symbol_path.suffix.lower() == ".xlsx":
        return {}

    return load_company_names(symbol_path)


def resolve_symbols(args):
    if args.mode == "1":
        symbols = load_symbols_from_file(args.csv_file)
    else:
        symbols = parse_symbol_items(args.manual_symbols)
        if not symbols:
            symbols = parse_symbol_items(MANUAL_SYMBOLS)

    if args.start_after:
        start_after = normalize_symbol(args.start_after)
        if start_after in symbols:
            symbols = symbols[symbols.index(start_after) + 1 :]
        else:
            print(f"Warning: --start-after symbol not found: {start_after}")

    return symbols


def is_stale(path, max_age_days):
    if not path.exists():
        return True

    if max_age_days < 0:
        return False

    modified = datetime.fromtimestamp(path.stat().st_mtime)
    return modified < datetime.now() - timedelta(days=max_age_days)


def ensure_ticker_dir(data_dir, symbol):
    ticker_dir = data_dir / symbol
    ticker_dir.mkdir(parents=True, exist_ok=True)
    return ticker_dir


def make_nse_session(timeout):
    session = requests.Session()
    try:
        session.get("https://www.nseindia.com", headers=NSE_HEADERS, timeout=timeout)
    except Exception as exc:
        print(f"NSE session warm-up failed, continuing with per-symbol retries: {exc}")
    return session


def fetch_annual_report_rows(session, symbol, timeout):
    url = (
        "https://www.nseindia.com/api/annual-reports"
        f"?index=equities&symbol={quote(symbol, safe='')}"
    )
    response = session.get(url, headers=NSE_HEADERS, timeout=timeout)
    response.raise_for_status()
    return response.json().get("data", [])


def annual_report_filename(row):
    return f"{row['fromYr']}_{row['toYr']}.pdf"


def update_annual_reports(session, ticker_dir, symbol, max_reports, timeout):
    target_dir = ticker_dir / "annual_reports"
    target_dir.mkdir(parents=True, exist_ok=True)

    rows = fetch_annual_report_rows(session, symbol, timeout)[:max_reports]

    if not rows:
        print(f"{symbol}: no annual reports found from NSE")
        return False

    updated = False

    for row in rows:
        filename = annual_report_filename(row)
        target_file = target_dir / filename

        if target_file.exists() and target_file.stat().st_size > 1000:
            continue

        pdf_url = row["fileName"]
        response = session.get(pdf_url, headers=NSE_HEADERS, timeout=timeout)
        response.raise_for_status()

        target_file.write_bytes(response.content)
        updated = True
        print(f"{symbol}: downloaded annual report {filename}")

    return updated


def download_screener_page(session, ticker_dir, symbol, timeout):
    url = f"https://www.screener.in/company/{quote(symbol, safe='')}/consolidated/"
    response = session.get(url, headers=SCREENER_HEADERS, timeout=timeout)
    response.raise_for_status()

    target_dir = ticker_dir / "screener_finance"
    target_dir.mkdir(parents=True, exist_ok=True)

    target_file = target_dir / "company_page.html"
    old_text = target_file.read_text(encoding="utf-8", errors="ignore") if target_file.exists() else ""
    response_text = response.text

    target_file.write_text(response_text, encoding="utf-8")

    changed = old_text != response_text
    print(f"{symbol}: screener page refreshed")
    return changed or not old_text


def load_existing_transcript_urls(ticker_dir):
    metadata_path = ticker_dir / "concalls" / "screener_transcripts.json"

    if not metadata_path.exists():
        return []

    try:
        rows = json.loads(metadata_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []

    return [row.get("url") for row in rows if row.get("url")]


def update_transcripts(session, ticker_dir, symbol, max_transcripts):
    html = load_screener_html(ticker_dir)

    if not html:
        print(f"{symbol}: screener page missing, cannot parse transcripts")
        return False

    transcripts = extract_transcripts(html, max_transcripts)
    desired_urls = [item["url"] for item in transcripts]
    existing_urls = load_existing_transcript_urls(ticker_dir)
    existing_pdf_count = len(list((ticker_dir / "concalls").glob("*.pdf")))

    if desired_urls == existing_urls and existing_pdf_count >= min(len(desired_urls), max_transcripts):
        return False

    clear_old_concall_dirs(ticker_dir)
    target_dir = ticker_dir / "concalls"
    target_dir.mkdir(parents=True, exist_ok=True)

    metadata = []
    downloaded = 0
    failed = 0

    for index, item in enumerate(transcripts, start=1):
        filename = filename_from_url(item["url"], index, item["period"])
        output_path = target_dir / filename
        row = {
            "period": item["period"],
            "url": item["url"],
            "file": str(output_path.relative_to(ticker_dir)),
        }

        try:
            download_transcript_pdf(session, item["url"], output_path)
            downloaded += 1
        except Exception as exc:
            failed += 1
            row["error"] = str(exc)
            print(f"{symbol}: transcript failed {item['period'] or index}: {exc}")

        metadata.append(row)

    save_transcript_metadata(ticker_dir, metadata)
    print(f"{symbol}: transcripts found={len(transcripts)} downloaded={downloaded} failed={failed}")
    return True


def read_news_links(ticker_dir):
    links_path = ticker_dir / "news" / "news_links.txt"

    if not links_path.exists():
        return []

    return [line.strip() for line in links_path.read_text(encoding="utf-8", errors="ignore").splitlines() if line.strip()]


def update_news_links(ticker_dir, symbol, company_name, max_links, timeout, use_fallback):
    query = build_query(symbol, company_name)
    items = get_news_items(query, max_links, timeout=timeout, use_fallback=use_fallback)
    new_links = [item["link"] for item in items]
    old_links = read_news_links(ticker_dir)

    write_news_outputs(ticker_dir / "news", items)

    changed = new_links != old_links
    print(f"{symbol}: news links saved={len(items)}")
    return changed or len(old_links) == 0


def has_minimum_data(ticker_dir):
    annual_reports = list((ticker_dir / "annual_reports").glob("*.pdf"))
    screener_page = ticker_dir / "screener_finance" / "company_page.html"
    news_links = ticker_dir / "news" / "news_links.txt"
    transcript_metadata = ticker_dir / "concalls" / "screener_transcripts.json"

    return {
        "annual_reports": bool(annual_reports),
        "screener_page": screener_page.exists(),
        "news_links": news_links.exists() and len(read_news_links(ticker_dir)) > 0,
        "transcripts_metadata": transcript_metadata.exists(),
    }


def analysis_needs_refresh(ticker_dir, data_updated, force_analysis):
    report_path = ticker_dir / "investment_analysis.md"

    if force_analysis:
        return True

    if not report_path.exists():
        return True

    if data_updated:
        return True

    return False


def print_score_log(ticker_dir):
    report_path = ticker_dir / "investment_analysis.md"

    if not report_path.exists():
        print(f"{ticker_dir.name}: score log unavailable, investment_analysis.md missing")
        return

    try:
        scores, missing = parse_scores(report_path)
    except Exception as exc:
        print(f"{ticker_dir.name}: score log unavailable, could not parse final summary table: {exc}")
        return

    print(f"{ticker_dir.name}: final summary scores")
    print("| Stock | Business Quality | Management | Financial Strength | Growth | Governance | Valuation |")
    print("|---|---:|---:|---:|---:|---:|---:|")
    print(
        "| "
        f"{ticker_dir.name} | "
        f"{format_score(scores['business']) if 'business' in scores else 'NA'} | "
        f"{format_score(scores['management']) if 'management' in scores else 'NA'} | "
        f"{format_score(scores['financial']) if 'financial' in scores else 'NA'} | "
        f"{format_score(scores['growth']) if 'growth' in scores else 'NA'} | "
        f"{format_score(scores['governance']) if 'governance' in scores else 'NA'} | "
        f"{format_score(scores['valuation']) if 'valuation' in scores else 'NA'} |"
    )

    if missing:
        missing_labels = ", ".join(SCORE_METRICS[name] for name in missing)
        print(f"{ticker_dir.name}: score log missing metrics: {missing_labels}")


def run_ai_analysis(ticker_dir, args, prompt):
    evidence = build_evidence_pack(
        ticker_dir=ticker_dir,
        total_limit=args.evidence_chars,
        per_file_limit=args.per_file_chars,
    )

    if args.save_evidence:
        (ticker_dir / "analysis_evidence_pack.md").write_text(evidence, encoding="utf-8")

    if args.dry_run:
        print(f"{ticker_dir.name}: dry-run, analysis not generated")
        return

    report = call_codex_cli(
        prompt=prompt,
        evidence=evidence,
        model=args.model,
        timeout=args.codex_timeout,
    )

    write_report(
        output_path=ticker_dir / "investment_analysis.md",
        ticker=ticker_dir.name,
        model=args.model,
        evidence_chars=len(evidence),
        report_text=report,
    )

    print(f"{ticker_dir.name}: investment_analysis.md generated")
    print_score_log(ticker_dir)


def process_symbol(symbol, args, context):
    ticker_dir = ensure_ticker_dir(Path(args.data_dir), symbol)
    data_updated = False

    print("\n" + "=" * 70)
    print(f"Processing {symbol}")
    print("=" * 70)

    if args.skip_data_refresh:
        print(f"{symbol}: data refresh skipped by --skip-data-refresh")
    else:
        try:
            annual_updated = update_annual_reports(
                context["nse_session"],
                ticker_dir,
                symbol,
                args.max_annual_reports,
                args.request_timeout,
            )
            data_updated = data_updated or annual_updated
        except Exception as exc:
            print(f"{symbol}: annual report update failed: {exc}")

        screener_path = ticker_dir / "screener_finance" / "company_page.html"
        if args.force_data or is_stale(screener_path, args.screener_max_age_days):
            try:
                screener_updated = download_screener_page(
                    context["http_session"],
                    ticker_dir,
                    symbol,
                    args.request_timeout,
                )
                data_updated = data_updated or screener_updated
            except Exception as exc:
                print(f"{symbol}: screener page update failed: {exc}")

        if args.skip_transcripts:
            print(f"{symbol}: transcript update skipped by --skip-transcripts")
        else:
            try:
                transcript_updated = update_transcripts(
                    context["http_session"],
                    ticker_dir,
                    symbol,
                    args.max_transcripts,
                )
                data_updated = data_updated or transcript_updated
            except Exception as exc:
                print(f"{symbol}: transcript update failed: {exc}")

        pre_news_status = has_minimum_data(ticker_dir)
        core_data_missing = any(
            not pre_news_status[name]
            for name in ("annual_reports", "screener_page", "transcripts_metadata")
        )

        news_path = ticker_dir / "news" / "news_links.txt"
        if args.skip_news:
            print(f"{symbol}: news update skipped by --skip-news")
        elif args.skip_analysis_if_missing_data and core_data_missing:
            print(f"{symbol}: news update skipped because core data sections are missing")
        elif args.force_data or is_stale(news_path, args.news_max_age_days):
            try:
                company_name = context["company_names"].get(symbol, "")
                news_updated = update_news_links(
                    ticker_dir,
                    symbol,
                    company_name,
                    args.max_news_links,
                    args.news_timeout,
                    not args.disable_news_fallback,
                )
                data_updated = data_updated or news_updated
            except Exception as exc:
                print(f"{symbol}: news update failed: {exc}")

    status = has_minimum_data(ticker_dir)
    missing = [name for name, present in status.items() if not present]

    if missing:
        print(f"{symbol}: warning, missing data sections: {', '.join(missing)}")
        if args.skip_analysis_if_missing_data:
            report_path = ticker_dir / "investment_analysis.md"
            if report_path.exists() and not data_updated and not args.force_analysis:
                print(f"{symbol}: best condition, data current and investment_analysis.md exists")
                print_score_log(ticker_dir)
                return {"symbol": symbol, "data_updated": data_updated, "analysis": "already_current"}

            print(f"{symbol}: analysis skipped because required data sections are missing")
            return {"symbol": symbol, "data_updated": data_updated, "analysis": "skipped"}

    if args.skip_analysis:
        print(f"{symbol}: analysis skipped by --skip-analysis")
        return {"symbol": symbol, "data_updated": data_updated, "analysis": "skipped"}

    if analysis_needs_refresh(ticker_dir, data_updated, args.force_analysis):
        try:
            run_ai_analysis(ticker_dir, args, context["prompt"])
            return {"symbol": symbol, "data_updated": data_updated, "analysis": "generated"}
        except Exception as exc:
            print(f"{symbol}: AI analysis failed: {exc}")
            return {"symbol": symbol, "data_updated": data_updated, "analysis": "failed"}

    print(f"{symbol}: best condition, data current and investment_analysis.md exists")
    print_score_log(ticker_dir)
    return {"symbol": symbol, "data_updated": data_updated, "analysis": "already_current"}


def main():
    parser = argparse.ArgumentParser(
        description="One-stop stock data refresh and Codex analysis pipeline."
    )
    parser.add_argument(
        "--mode",
        choices=["1", "2"],
        help="1 = read symbols from CSV, 2 = use manual symbol list.",
    )
    parser.add_argument("--csv-file", default=str(DEFAULT_CSV_FILE))
    parser.add_argument("--manual-symbols", nargs="*", help="Mode 2 override, comma-separated or space-separated.")
    parser.add_argument("--start-after", help="Resume after this normalized symbol in the selected symbol list.")
    parser.add_argument("--data-dir", default=str(DATA_DIR))
    parser.add_argument("--max-annual-reports", type=int, default=MAX_ANNUAL_REPORTS)
    parser.add_argument("--max-transcripts", type=int, default=MAX_TRANSCRIPTS)
    parser.add_argument("--max-news-links", type=int, default=MAX_NEWS_LINKS)
    parser.add_argument("--skip-transcripts", action="store_true", help="Do not refresh/download concall transcripts.")
    parser.add_argument("--screener-max-age-days", type=int, default=7)
    parser.add_argument("--news-max-age-days", type=int, default=1)
    parser.add_argument("--force-data", action="store_true")
    parser.add_argument("--skip-data-refresh", action="store_true", help="Use only existing local data and reports.")
    parser.add_argument("--force-analysis", action="store_true")
    parser.add_argument("--skip-analysis", action="store_true")
    parser.add_argument(
        "--skip-analysis-if-missing-data",
        action="store_true",
        help="Do not generate AI analysis when required data sections are missing.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Refresh/audit data and build evidence, but do not call Codex.")
    parser.add_argument("--save-evidence", action="store_true")
    parser.add_argument("--model", default="")
    parser.add_argument("--evidence-chars", type=int, default=220000)
    parser.add_argument("--per-file-chars", type=int, default=25000)
    parser.add_argument("--codex-timeout", type=int, default=1800)
    parser.add_argument("--request-timeout", type=int, default=60)
    parser.add_argument("--news-timeout", type=int, default=60)
    parser.add_argument("--disable-news-fallback", action="store_true")
    parser.add_argument("--skip-news", action="store_true", help="Do not refresh Google/fallback news links.")
    parser.add_argument("--sleep", type=float, default=0.5)
    args = parser.parse_args()

    if not args.mode:
        args.mode = input("Enter mode (1 = CSV, 2 = manual list): ").strip()

    if args.mode not in {"1", "2"}:
        raise ValueError("Mode must be 1 or 2.")

    symbols = resolve_symbols(args)

    if not symbols:
        raise ValueError("No symbols found. Add MANUAL_SYMBOLS or pass --manual-symbols.")

    prompt_path = PROMPT_FILE
    if not prompt_path.exists() and not args.skip_analysis:
        raise FileNotFoundError(f"Prompt file not found: {prompt_path}")

    context = {
        "nse_session": make_nse_session(args.request_timeout),
        "http_session": requests.Session(),
        "company_names": load_company_names_from_file(args.csv_file),
        "prompt": prompt_path.read_text(encoding="utf-8") if prompt_path.exists() else "",
    }

    counts = {
        "symbols": len(symbols),
        "data_updated": 0,
        "analysis_generated": 0,
        "analysis_current": 0,
        "analysis_skipped": 0,
        "analysis_failed": 0,
    }

    for symbol in symbols:
        result = process_symbol(symbol, args, context)

        if result["data_updated"]:
            counts["data_updated"] += 1

        analysis_status = result["analysis"]
        if analysis_status == "generated":
            counts["analysis_generated"] += 1
        elif analysis_status == "already_current":
            counts["analysis_current"] += 1
        elif analysis_status == "skipped":
            counts["analysis_skipped"] += 1
        elif analysis_status == "failed":
            counts["analysis_failed"] += 1

        if args.sleep > 0:
            time.sleep(args.sleep)

    print("\nDONE")
    print(json.dumps(counts, indent=2))

    return 0 if counts["analysis_failed"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
